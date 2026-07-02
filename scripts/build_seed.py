#!/usr/bin/env python3
"""
build_seed.py — extract the three source spreadsheets into scripts/seed.json,
the input to import_seed.ts.

GOTCHA (do not relearn): open workbooks with data_only=True so that cached date
formulas in the sheets (e.g. Monitoring's `=E-90`, `=D+365`) resolve to real
dates instead of formula strings. Entity cells are slash-separated, never
comma-joined. All dates are emitted as local "yyyy-mm-dd".

Run:  python3 scripts/build_seed.py            # writes scripts/seed.json
      python3 scripts/build_seed.py --validate # also diff counts vs reference
"""
from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "source")
OUT = os.path.join(HERE, "seed.json")
REFERENCE = os.path.join(HERE, "seed.reference.json")
LOGO = os.path.join(HERE, "..", "web", "public", "logo.png")

APPROVED = {"Unassigned", "Mike Gregory", "Jack Griffin", "Harrison Fritz", "Intern"}


def local_iso(v):
    """Any cell value -> local 'yyyy-mm-dd' or None. Never shifts by timezone."""
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
    s = str(v).strip()
    if not s:
        return None
    # already an ISO-ish string
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            d = dt.datetime.strptime(s[:10], fmt)
            return f"{d.year:04d}-{d.month:02d}-{d.day:02d}"
        except ValueError:
            continue
    return None


def date_or_raw(v):
    """Local ISO if parseable, else the raw trimmed string (e.g. 'Q2'/'Q3'
    quarter placeholders used in Potential trips). import_seed nulls out any
    value that isn't a real date before writing the `date` column."""
    iso = local_iso(v)
    if iso:
        return iso
    return cell(v)


def cell(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def norm_level(raw):
    """Strip the 'BFO - ' prefix; normalize to 'Level N'."""
    if not raw:
        return "Level 1"
    s = str(raw).replace("BFO - ", "").strip()
    return s if s.startswith("Level") else "Level 1"


def load(name):
    return openpyxl.load_workbook(os.path.join(SRC, name), data_only=True)


# ─── travel ──────────────────────────────────────────────────────────────────
def extract_travel():
    ws = load("Travel_Schedule.xlsx").active
    section = "upcoming"
    marker = {"Upcoming": "upcoming", "Potential": "potential", "Archived": "archived"}
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header
            continue
        a = r[0]
        if a and isinstance(a, str) and any(k in a for k in marker):
            for k, v in marker.items():
                if k in a:
                    section = v
            continue
        # skip fully blank rows (a row may carry stray values past col 10)
        if not any(cell(r[c]) for c in (0, 2, 3, 5)):  # date, city, analyst, event
            continue
        rows.append({
            "section": section,
            "date": date_or_raw(r[0]),
            "days": cell(r[1]),
            "city": cell(r[2]),
            "analyst": cell(r[3]),
            "monitoringVisits": cell(r[4]),
            "event": cell(r[5]),
            "flight": cell(r[6]),
            "hotel": cell(r[7]),
            "car": cell(r[8]),
            "notesOtherVisits": cell(r[9]),
        })
    return rows


# ─── monitoring ────────────────────────────────────────────────────────────────
def extract_monitoring():
    ws = load("Monitoring.xlsx").active
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if r[0] is None and r[1] is None:
            continue
        rows.append({
            "analyst": cell(r[0]),
            "fund": cell(r[1]),
            "level": r[2],  # raw; normalized on import/build
            "mostRecent": local_iso(r[3]),
            "monitoringDate": local_iso(r[4]),
        })
    return rows


# ─── PRC archive ───────────────────────────────────────────────────────────────
def extract_prc_archive():
    ws = load("PRC.xlsx")["Meeting Archive"]
    rows = []
    header_seen = False
    for r in ws.iter_rows(values_only=True):
        if not header_seen:
            if r[0] == "Date":
                header_seen = True
            continue
        if all(c is None for c in r[:7]):
            continue
        rows.append({
            "meetingDate": local_iso(r[0]),
            "macro": cell(r[1]),
            "presentation": cell(r[2]),
            "act40": cell(r[3]),
            "hedgeFund": cell(r[4]),
            "private": cell(r[5]),
            "newFunds": cell(r[6]),
        })
    return rows


# ─── PRC mapping grid (Meeting Schedule sheet) ───────────────────────────────
def extract_prc_mapping():
    """3 (Presentation, Entity) column-pairs: 40-Act, Hedge Fund, Private.
    Group entities by presentation -> map40 / mapHF / mapPriv, plus global lists.
    Entity cells stay slash-joined (e.g. 'Westwood/Aspenleaf')."""
    ws = load("PRC.xlsx")["Meeting Schedule"]
    pairs = {"act40": (0, 1), "hf": (2, 3), "priv": (4, 5)}
    maps = {"act40": {}, "hf": {}, "priv": {}}
    globals_ = {"act40": [], "hf": [], "priv": []}
    presentations = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        for cat, (pc, ec) in pairs.items():
            pres = cell(r[pc]) if pc < len(r) else None
            ent = cell(r[ec]) if ec < len(r) else None
            if pres == "Flex":  # Flex is derived at runtime, never a schedule row
                continue
            if pres and pres not in presentations:
                presentations.append(pres)
            if pres and ent:
                maps[cat].setdefault(pres, [])
                if ent not in maps[cat][pres]:
                    maps[cat][pres].append(ent)
                if ent not in globals_[cat]:
                    globals_[cat].append(ent)
    to_global = lambda lst: [{"name": n, "flex": False} for n in lst]
    return {
        "presentations": presentations,
        "act40Global": to_global(globals_["act40"]),
        "hedgeFundGlobal": to_global(globals_["hf"]),
        "privateGlobal": to_global(globals_["priv"]),
        "map40": maps["act40"],
        "mapHF": maps["hf"],
        "mapPriv": maps["priv"],
        # Flex membership is derived at runtime by the PRC module (oldest archive
        # appearance). Seed leaves these empty; the app computes them.
        "flexPriv": [],
        "flexHF": [],
    }


def derive_prc_schedule(archive, mapping):
    """One schedule row per presentation. most_recent = latest archive meeting
    for that presentation; projected_next is left null (app computes it)."""
    latest = {}
    for a in archive:
        p = a.get("presentation")
        d = a.get("meetingDate")
        if p and d and (p not in latest or d > latest[p]):
            latest[p] = d
    rows = []
    for pres in mapping["presentations"]:
        rows.append({
            "presentation": pres,
            "mostRecent": latest.get(pres),
            "projectedNext": None,
            "macro": "",
            "act40": "/".join(mapping["map40"].get(pres, [])),
            "hedgeFund": "/".join(mapping["mapHF"].get(pres, [])),
            "private": "/".join(mapping["mapPriv"].get(pres, [])),
            "newFunds": "",
        })
    return rows


def load_logo():
    try:
        with open(LOGO, "rb") as f:
            return "data:image/png;base64," + base64.b64encode(f.read()).decode()
    except OSError:
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--validate", action="store_true", help="diff counts vs seed.reference.json")
    args = ap.parse_args()

    travel = extract_travel()
    monitoring = extract_monitoring()
    archive = extract_prc_archive()
    mapping = extract_prc_mapping()
    schedule = derive_prc_schedule(archive, mapping)

    seed = {
        "travel": travel,
        "monitoring": monitoring,
        "prcArchive": archive,
        "prcSchedule": schedule,
        "prcEntities": [],  # entities grid; populated in a later version
        "prcMapping": mapping,
        "logo": load_logo(),
    }
    with open(OUT, "w") as f:
        json.dump(seed, f, indent=2)
    print(f"wrote {OUT}")
    counts = {k: len(v) for k, v in seed.items() if isinstance(v, list)}
    print("counts:", counts)

    if args.validate and os.path.exists(REFERENCE):
        with open(REFERENCE) as f:
            ref = json.load(f)
        print("\nvalidation vs reference:")
        for k in ("travel", "monitoring", "prcArchive", "prcSchedule"):
            got, exp = len(seed.get(k, [])), len(ref.get(k, []))
            flag = "ok" if got == exp else "DIFF"
            print(f"  {k:14} got={got:<4} reference={exp:<4} [{flag}]")


if __name__ == "__main__":
    sys.exit(main())
